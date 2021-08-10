from django.shortcuts import render, redirect
from django.core.mail import EmailMessage
from django.conf import settings
from django.utils import timezone 

#exportar excel
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView

import os
from enviomail.models import *

def index(request,*args, **kwargs):
    lista_correos = RegistroEnvioMail.objects.all()
    contexto = {'ListMails':lista_correos}
    return render(request, 'index.html', contexto)

BASE_DIR_MAIL = 'C:/Users/WILMER/Documents/prueba'

def EnviarMail(factura, ruta):
    # contexto = {'mail':mail}
    # template = get_template('platilla_mail.html')
    # content = template.render(contexto)
    email = EmailMessage(
        '900540156; Fundación Clínica del Rio; ' + factura + '; 01; Fundación Clínica del Rio',#ENCABEZADO
        # content,#CUERPO DEL MENSAJE HECHO EN HTML
        'Buen día, correo para notificación electrónica: aux.contable@clinicadelrio.org',
        settings.EMAIL_HOST_USER,
        ['wildau@hotmail.com'],
        # cc = [] #copia de correo
    )
    # email.attach('platilla_mail.html',content, 'text/html')
    # email.content_subtype = "html"#PARA INTEGRAR UN HTML DENTRO DEL CUERPO DEL CORREO
    email.attach_file(ruta)#PARA ADJUNTAR EL ARCHIVO EN lA RUTA INDICADA
    email.send()

def EnviarMailAuto(request):
    #entrar a la carpeta raiz donde se guardan las carpetas con los nombres de los dias
    # with os.scandir(BASE_DIR_MAIL) as archivos:
    #     archivos = [archivo.name for archivo in archivos if archivo.is_dir()]
    archivos = ['file']
    # print(archivos)
    #recorrer cada una de las carpetas con nombre de los dias
    for corredor in archivos:
        BASE_DIR_MAIL_SECON = BASE_DIR_MAIL + '/' + corredor

        with os.scandir(BASE_DIR_MAIL_SECON) as FileFacturas:
            FileFactura = [FileFactura.name for FileFactura in FileFacturas if FileFactura.is_dir()]

        #recorrer la carpeta con nombre de las facturas
        for SeconCorredor in FileFactura:
            BASE_DIR_MAIL_THRE = BASE_DIR_MAIL_SECON + '/' + SeconCorredor

            # extraer el archivo .zip para enviar por correo
            with os.scandir(BASE_DIR_MAIL_THRE) as FileZips:
                FileZips = [FileZip.name for FileZip in FileZips if FileZip.is_file() and FileZip.name.endswith('.zip')] #ARCHIVO EXTRAIDO
                RutaFinal = BASE_DIR_MAIL_THRE + '/'+ str(FileZips[0])

            #validar si el archivo actual ha sido enviado anteriormente
            CorreoEnviados = RegistroEnvioMail.objects.all()
            cont = 0
            for CorreoEnviado in CorreoEnviados:
                if str(FileZips[0]) == CorreoEnviado.archivo:
                    cont += 1

            # enviar archivo .zip     
            if cont == 0:
                EnviarMail(SeconCorredor,RutaFinal)

                #Registrar el correo enviado
                registroEnvio = RegistroEnvioMail()
                registroEnvio.numero_factura = SeconCorredor
                registroEnvio.fecha_envio = timezone.now()
                registroEnvio.Estado = 'ENVIADO'
                registroEnvio.ruta = RutaFinal
                registroEnvio.archivo = FileZips[0]
                registroEnvio.save()
                print('Nuevo Registro, la Factura # : ', SeconCorredor, ' ha sido Enviada !', 'día: ',corredor)
            else:
                print('La Factura #: ', SeconCorredor, 'Se habia Enviado con Anterioridad !', 'día: ',corredor)

    return redirect('home')
    


#reporte excel emailo enviados
class Reporte_SendMail_excel(TemplateView):
    def get(self, request, *args, **kwargs):
        mails = RegistroEnvioMail.objects.all()
        wb = Workbook()
        ws = wb.active
        # ws['A1'] = 'REGISTRO DE CORREOS ENVIADOS A MUTUAL SER EPS - FACTURACIÓN ELECTRÓNICA'
        ws['A1']= 'ID'
        ws['B1']= 'Numero de Factura'
        ws['C1'] = 'Fecha de Envio'
        ws['D1'] = 'Estado'
        ws['E1'] = 'Nombre del Archivo'
        ws['f1'] = 'Ruta' 

        cont = 2
        for mail in mails:
            ws.cell(row=cont, column = 1).value = mail.id
            ws.cell(row=cont, column = 2).value = mail.numero_factura
            ws.cell(row=cont, column = 3).value = str(mail.fecha_envio)
            ws.cell(row=cont, column = 4).value = mail.Estado
            ws.cell(row=cont, column = 5).value = mail.archivo
            ws.cell(row=cont, column = 6).value = mail.ruta
            cont +=1
        
        nombre_reporte = "Reporte_SendMail_excel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_reporte)
        response['content-Disposition'] = content
        wb.save(response)
        return response
#==================================================================================================
